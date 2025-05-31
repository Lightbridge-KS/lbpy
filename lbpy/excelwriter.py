import logging
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
import os

__all__ = ["ExcelTemplateWriter"]


class ExcelTemplateWriter:
    """Excel template writer for structured data insertion with incremental editing support."""

    def __init__(self, template_path, sheet_name=None):
        """
        Initialize Excel template writer with template file.

        Parameters:
        - template_path: Path to the Excel template file
        - sheet_name: Optional sheet name (uses active sheet if None)
        """
        self.template_path = template_path
        self.sheet_name = sheet_name

        # Validate template file exists during initialization
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Template file not found: {template_path}")

        try:
            # Load the existing workbook (preserves formatting)
            self.workbook = load_workbook(self.template_path)

            # Get the worksheet reference
            if self.sheet_name:
                if self.sheet_name not in self.workbook.sheetnames:
                    raise ValueError(f"Sheet '{self.sheet_name}' not found in workbook")
                self.worksheet = self.workbook[self.sheet_name]
            else:
                self.worksheet = self.workbook.active

        except InvalidFileException:
            logging.error(f"Invalid Excel file: {self.template_path}")
            raise
        except Exception:
            logging.exception("Unexpected error occurred while loading template")
            raise

    def __repr__(self):
        """
        Return technical string representation of ExcelTemplateWriter instance.

        Returns
        -------
        str
            Technical representation showing class name and key attributes
        """
        return f"ExcelTemplateWriter(template_path='{self.template_path}', sheet_name={self.sheet_name!r})"

    def __str__(self):
        """
        Return user-friendly string representation of ExcelTemplateWriter instance.

        Returns
        -------
        str
            Human-readable description of the writer instance
        """
        sheet_info = f"sheet '{self.sheet_name}'" if self.sheet_name else "active sheet"
        return f"ExcelTemplateWriter for {os.path.basename(self.template_path)} ({sheet_info})"

    def modify(self, data, column_mapping, start_row):
        """
        Modify workbook with data using specified column mapping and start row.

        Parameters:
        - data: Dictionary or list of dictionaries containing data to write
        - column_mapping: Dictionary mapping data keys to Excel columns
        - start_row: Starting row for data insertion

        Returns:
        - self: For method chaining
        """

        try:
            # Handle both single dictionary and list of dictionaries
            if isinstance(data, dict):
                data = [data]

            # Write data to the table
            for row_offset, record in enumerate(data):
                current_row = start_row + row_offset

                for field, column in column_mapping.items():
                    if field in record:
                        cell_address = f"{column}{current_row}"
                        self.worksheet[cell_address] = record[field]

            logging.info(
                f"Successfully modified {len(data)} records starting from row {start_row}"
            )

        except Exception:
            logging.exception("Unexpected error occurred while modifying workbook")
            raise

        return self  # Enable method chaining

    def modify_safe(self, data, column_mapping, start_row):
        """
        Modify workbook with safe handling of missing dictionary keys.

        Parameters:
        - data: Dictionary or list of dictionaries containing data to write
        - column_mapping: Dictionary mapping data keys to Excel columns
        - start_row: Starting row for data insertion

        Returns:
        - self: For method chaining
        """

        try:
            # Handle both single dict and list of dicts
            if isinstance(data, dict):
                data = [data]

            records_written = 0
            for row_offset, record in enumerate(data):
                current_row = start_row + row_offset

                # Write each field, using empty string for missing keys
                for field, column in column_mapping.items():
                    cell_address = f"{column}{current_row}"
                    self.worksheet[cell_address] = record.get(field, "")

                records_written += 1

            logging.info(
                f"Successfully modified {records_written} records starting from row {start_row}"
            )

        except Exception:
            logging.exception("Error occurred while modifying workbook")
            raise

        return self  # Enable method chaining

    def write(self, output_path):
        """
        Write the modified workbook to output path.

        Parameters:
        - output_path: Path where the modified workbook will be saved

        Returns:
        - None
        """

        try:
            self.workbook.save(output_path)
            print(f"Successfully saved workbook to {output_path}")

        except PermissionError:
            logging.error(
                f"Permission denied when saving to {output_path}. File might be open in Excel."
            )
            raise
        except Exception:
            logging.exception("Unexpected error occurred while saving workbook")
            raise

        return None

    def validate_data(self, data, column_mapping):
        """
        Validate that data contains expected keys from column mapping.

        Parameters:
        - data: Dictionary or list of dictionaries to validate
        - column_mapping: Dictionary mapping data keys to Excel columns

        Returns:
        - bool: True if all required keys are present, False otherwise
        """
        if isinstance(data, dict):
            data = [data]

        missing_keys = []
        for i, record in enumerate(data):
            missing = [key for key in column_mapping.keys() if key not in record]
            if missing:
                missing_keys.append(f"Record {i}: {missing}")

        if missing_keys:
            logging.warning(f"Missing keys found: {missing_keys}")

        return len(missing_keys) == 0

    def reset_workbook(self):
        """
        Reset workbook to original template state.

        Returns:
        - self: For method chaining
        """
        try:
            self.workbook = load_workbook(self.template_path)

            # Get the worksheet reference
            if self.sheet_name:
                self.worksheet = self.workbook[self.sheet_name]
            else:
                self.worksheet = self.workbook.active

            logging.info("Workbook reset to original template state")

        except Exception:
            logging.exception("Error occurred while resetting workbook")
            raise

        return self  # Enable method chaining
